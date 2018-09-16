import sys
import os
# import xlrd
# import xlwt
# import xlsxwriter
# from pathlib import Path
import pandas as pd
from openpyxl import load_workbook


def does_file_exist(usr_dir, new_file, all_data_file):

    found_new_file = False
    found_all_data_file = False

    #check for files in directory
    if os.path.isdir(usr_dir):                                     #check if directory exists
        for root, dirs, files in os.walk(usr_dir):
            for file in files:
                if file.endswith(new_file):
                    found_new_file = True
                if file.endswith(all_data_file):
                    found_all_data_file = True
    else:                                                          #directory does NOT exist, exit program
        print('Directory not found')
        quit()

    #check if both files have been found
    if(found_new_file == True and found_all_data_file == True):     #both files found
        return True
    else:                                                           #files NOT found
        return False






def update_files(usr_dir, new_file, all_data_file):

    location_new_file = (usr_dir + '/' + new_file)                                          #file with new data
    location_all_data_file = (usr_dir + '/' + all_data_file)                                #file with all data

    xl = pd.ExcelFile(location_new_file)                                                    #get location of new_data file

    data = xl.parse('Sheet1')                                                               #get the data from the new_file without header

    writer = pd.ExcelWriter(location_all_data_file, engine='xlsxwriter')                    #specify the writer

    data.to_excel(writer, sheet_name='Sheet1', startrow=0, startcol=-1, header=None)        #write data to the new file, with no header
                                                                                            #startrow=0 + header=None -> removes header
                                                                                            #startcol=-1 -> removes column that keeps counts

    append_df_to_excel(location_all_data_file, data, header=None, index=False)



#Helper function borrowed from: https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def main():

    found_file = False

    usr_dir = sys.argv[1]
    new_data_file = sys.argv[2]     #get name of new file
    all_data_file = sys.argv[3]     #get name of all data file

    found_file = does_file_exist(usr_dir, new_data_file, all_data_file)     #check if the input files exist

    if(found_file == False):         #if either files does NOT exist, print statement and exit
        print("Files not found")
        print("User Dir: " + usr_dir)
        print("New File: " + new_data_file)
        print("All File: " + all_data_file)
        quit()

    update_files(usr_dir, new_data_file, all_data_file)

    return


main()


#Resources
# - https://www.datacamp.com/community/tutorials/python-excel-tutorial
# - https://xlsxwriter.readthedocs.io/working_with_pandas.html (remove the header from data)
# - https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas